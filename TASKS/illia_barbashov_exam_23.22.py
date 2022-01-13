from openpyxl import load_workbook


class RecordIterator:
    """
    Класс реалізує ітератор для читання рядків з ексель таблиці і формує словник з кожного рядка
    """
    def __init__(self, file_name, sheet = "Sheet1", useActiveSheet=True):
        self.wb = load_workbook('sheet_sample.xlsx')
        if not useActiveSheet:
            self.ws = self.wb[sheet]
        else:
            self.ws = self.wb.active
        # read header
        self.header = next(ws.values)
        self.generator = ws.values

    def __next__(self):

        values = next(self.generator)
        return dict(zip(self.header, values))

    def __iter__(self):
        return self

if __name__ == '__main__':
    xlsx_file = input("Input excel file name")

    if not xlsx_file:
        xlsx_file = 'sheet_sample.xlsx'

    wb2 = load_workbook(xlsx_file)
    ws = wb2.active

    # Проста ітерація через ексель таблицю
    header = next(ws.values)
    for values in ws.values:
        print(dict(zip(header, values)))
    print("--------------------------------------------------------------")
    print("                Iterator                                       ")

    # Реалізація класу Ітератора
    recordIterator = RecordIterator("sheet_sample.xlsx")
    for record in recordIterator:
        print(record)


